import time

import pytest
from openpyxl import load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from datetime import datetime


@pytest.mark.usefixtures("setup")
class BaseClass:
    pass


def convert_excel_data(excel_value, index):
    """Converts Excel data to match the format of the Web data."""
    if index == 0 or index == 8:  # Convert float-like numbers to integers or handle special case
        if excel_value == 0 and index == 8:
            print("Converting '0' to '-'")  # Debugging line
            return "-"  # Special case for index 8
        try:
            return str(int(float(excel_value)))  # Convert to integer-like string, e.g., '55.0' -> '55'
        except ValueError:
            return excel_value  # Return as is if conversion fails

    elif index == 2:  # Convert 'true'/'false' to 'active'/'yes'
        if isinstance(excel_value, bool):
            return "active" if excel_value else "inactive"
        return "active" if excel_value.lower() in ["true"] else "inactive"

    elif index == 9: # Convert 'true'/'false' to 'Yes'/'No'
        if isinstance(excel_value, bool):
            return "yes" if excel_value else "no"
        return "yes" if excel_value.lower() == 'True' else "no"

    elif index == 3 or index == 6:  # Convert date formats
        try:
            return datetime.strptime(excel_value, "%d-%m-%Y").strftime("%d-%b-%Y").lower()
        except ValueError:
            return excel_value  # Return as is if the date is in an unexpected format

    elif index == 4 or index == 7:  # Convert 'none' values to '-'
        if excel_value is None:
            return "-"
        return "-" if excel_value.lower() == "none" else datetime.strptime(excel_value, "%d-%m-%Y").strftime("%d-%b-%Y").lower()

    return excel_value  # Return the value unchanged if no conversion is needed


def read_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    max_rows = 12  # Define the maximum number of rows to read
    #for i, row in enumerate(sheet.iter_rows(min_row=5, values_only=True)):
    for row in sheet.iter_rows(min_row=5, values_only=True):
        # if i >= max_rows:  # Stop after reading 10 rows
        #     break
        # Apply conversion to each cell in the row
        # Apply conversion to each cell in the row
        modified_row = [convert_excel_data(cell, index) for index, cell in enumerate(row)]
        data.append(modified_row)

    return data


# def convert_web_data(web_value, index):
#     """Converts web data to match the format of the Excel data."""
#     if index == 0 or index == 8:  # Convert float-like numbers
#         return f"{float(web_value)}"  # Ensure it matches Excel's float format
#
#     elif index == 2 or index == 9:  # Convert status and boolean values
#         return "true" if web_value.lower() in ["active", "yes"] else "false"
#
#     elif index == 3 or index == 6:  # Convert date formats
#         try:
#             return datetime.strptime(web_value, "%d-%b-%Y").strftime("%d-%m-%Y")
#         except ValueError:
#             return web_value  # Return as is if the date is in an unexpected format
#
#     elif index == 4 or index == 7:  # Convert 'none' values
#         return "none" if web_value == "-" else web_value
#
#     return web_value  # Return the value unchanged if no conversion is needed


def get_web_data(driver):
    wait = WebDriverWait(driver, 20)
    data = []

    while True:
        try:
            # Locate all rows in the table
            rows = wait.until(EC.presence_of_all_elements_located(
                (By.XPATH, "//tbody[@class='table-body ng-star-inserted']/tr")
            ))
            print(f"Found {len(rows)} rows on the current page")

            # Extract data from each row
            page_data = []
            for row in rows[0:]:  # Skipping the header row
                cols = row.find_elements(By.TAG_NAME, "td")
                row_data = [col.text.strip() for col in cols[1:]]  # Adjust index as per your table structure
                page_data.append(row_data)

            data.extend(page_data)
            print(f"Data from current page: {page_data}")

            # Attempt to locate and click the "Next" button
            next_button = driver.find_elements(By.XPATH, "//a[@aria-label='Next page' and contains(text(),'Next')]")
            if next_button:
                actions = ActionChains(driver)
                actions.move_to_element(next_button[0]).perform()
                print("Scrolled to the Next button")

                # Using JavaScript to click the button
                driver.execute_script("arguments[0].click();", next_button[0])
                print("Clicked Next button using JavaScript")

                # Wait for the rows to become stale (i.e., the page changes)
                wait.until(EC.staleness_of(rows[0]))
                print("Old page rows became stale, waiting for new rows to load")

                # # Wait for new rows to load
                # rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//table//tr")))
                # print(f"New rows are present: {len(rows)}")

                # Optional: Add a slight delay to ensure data is fully loaded
                time.sleep(2)  # Adjust this delay based on your page load speed
            else:
                print("No 'Next' button found, ending pagination")
                break

        except Exception as e:
            print(f"An error occurred: {e}")
            break

    return data


def normalize_data(data):
    return [[str(cell).strip().lower() for cell in row] for row in data]



